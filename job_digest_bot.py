from __future__ import annotations

import datetime as dt
import html
import json
import os
import re
import smtplib
import sys
from dataclasses import dataclass
from email.message import EmailMessage
from pathlib import Path
from typing import Iterable
from urllib.parse import quote_plus, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


PROFILE = {
    "name": "Vraj Patel",
    "email": "patelvraj09876@gmail.com",
    "experience": "2.5 years",
    "salary": "INR 14-16 LPA negotiable",
    "notice": "Immediate",
}

SMTP_ACCOUNT = "patelvraj09876@gmail.com"

TARGET_TITLES = [
    "AI/ML Engineer",
    "Machine Learning Engineer",
    "Generative AI Engineer",
    "LLM Engineer",
    "NLP Engineer",
    "Data Scientist",
    "Data Engineer",
]

KEYWORDS = [
    "python",
    "machine learning",
    "generative ai",
    "genai",
    "llm",
    "rag",
    "langchain",
    "langgraph",
    "nlp",
    "data engineer",
    "sql",
    "aws",
    "neo4j",
    "mlops",
]

PREFERRED_LOCATIONS = ["bangalore", "bengaluru", "pune", "hyderabad", "mumbai", "remote", "hybrid", "india"]
EXCLUDED_LOCATIONS = ["ahmedabad", "vadodara", "gurgaon", "gurugram", "noida"]

JOB_GROUPS = {
    "AI/ML": [
        "AI ML Engineer Python LLM",
        "Generative AI Engineer RAG Python",
        "Machine Learning Engineer Python",
        "LLM Engineer LangChain RAG",
        "NLP Engineer Python LLM",
    ],
    "Data Engineer": [
        "Data Engineer Python SQL",
        "Data Engineer ETL Python SQL",
        "Data Engineer AWS Python SQL",
        "Data Engineer Spark Python",
        "Data Engineer AI Python SQL",
    ],
    "Data Scientist": [
        "Data Scientist Python Machine Learning",
        "Data Scientist NLP Python",
        "Data Scientist LLM Python",
        "Data Scientist SQL Python",
        "Machine Learning Data Scientist Python",
    ],
}


@dataclass
class Job:
    company: str
    role: str
    location: str
    mode: str
    source: str
    link: str
    snippet: str
    score: int
    next_action: str


SEEDED_FALLBACK_JOBS = [
    Job("BayOne Solutions", "Artificial Intelligence Engineer", "India", "Remote", "LinkedIn", "https://www.linkedin.com/jobs/view/4416388565/", "Easy Apply; actively reviewing applicants. Good AI engineer title fit.", 88, "Open job and review Easy Apply form"),
    Job("Precision AQ", "AI/ML Engineer II", "Bengaluru", "Remote", "LinkedIn", "https://www.linkedin.com/jobs/view/4417508818/", "Easy Apply; strong title and remote mode.", 86, "Open job and review Easy Apply form"),
    Job("Chubb", "AI/ML Engineer", "Bengaluru", "Hybrid", "LinkedIn", "https://www.linkedin.com/jobs/view/4404058101/", "Easy Apply; actively reviewing applicants.", 84, "Open job and review Easy Apply form"),
    Job("Teradata", "AI Engineer", "Pune/Pimpri-Chinchwad", "Hybrid", "LinkedIn", "https://www.linkedin.com/jobs/view/4370700303/", "Easy Apply; actively reviewing applicants. Good Pune option.", 83, "Open job and review Easy Apply form"),
    Job("HackerRank", "Machine Learning Engineer, Integrity", "Bengaluru East", "Hybrid", "LinkedIn", "https://www.linkedin.com/jobs/view/4398830613/", "Strong ML engineering brand; likely competitive.", 81, "Review JD and tailor resume keywords"),
    Job("Verveo", "Generative AI Engineer", "Bengaluru", "Check listing", "LinkedIn/Public", "https://in.linkedin.com/jobs/view/generative-ai-engineer-at-verveo-4395975109", "RAG, vector DBs, LangChain/LlamaIndex/LangGraph, Python, PyTorch/TensorFlow, Docker, FastAPI/Flask.", 91, "Verify active listing and email hr@verveo.com"),
    Job("APPIT Software Solutions", "Generative AI Engineer", "Hyderabad", "Check listing", "Company Careers", "https://www.appitsoftware.com/careers/generative-ai-engineer-hyderabad", "GenAI, RAG, vector DBs, LangChain/OpenAI, FastAPI, Hugging Face.", 89, "Apply on company careers page"),
    Job("NeuLeap", "Generative AI Engineer", "Pune", "Check listing", "Company Careers", "https://neuleap.ai/careers/5", "RAG, prompt engineering, LangChain/LangGraph/LlamaIndex, Python, REST APIs, Docker/K8s.", 88, "Verify active listing and email admin@neuleap.ai"),
]


def norm(value: str) -> str:
    value = re.sub(r"<[^>]+>", " ", value or "")
    return re.sub(r"\s+", " ", html.unescape(value)).strip()


def is_remote(text: str) -> bool:
    return bool(re.search(r"\b(remote|work from home|wfh)\b", text, re.I))


def is_allowed_location(text: str) -> bool:
    low = text.lower()
    if any(x in low for x in EXCLUDED_LOCATIONS) and not is_remote(low):
        return False
    return any(x in low for x in PREFERRED_LOCATIONS)


def score_job(role: str, company: str, location: str, snippet: str) -> int:
    blob = f"{role} {company} {location} {snippet}".lower()
    score = 50
    if any(t.lower() in blob for t in TARGET_TITLES):
        score += 10
    for word in KEYWORDS:
        if word in blob:
            score += 3
    if re.search(r"\b(2|3|4)\+?\s*(years|yrs|yr)\b", blob):
        score += 8
    if "rag" in blob or "langchain" in blob or "llm" in blob:
        score += 8
    if is_remote(blob) or "hybrid" in blob:
        score += 4
    if any(x in blob for x in EXCLUDED_LOCATIONS) and not is_remote(blob):
        score -= 100
    return max(0, min(score, 100))


def job_group(job: Job) -> str:
    blob = f"{job.role} {job.snippet}".lower()
    if re.search(r"\b(data engineer|etl|data pipeline|data warehouse|spark|airflow|bigquery|databricks)\b", blob):
        return "Data Engineer"
    if re.search(r"\b(data scientist|statistical|analytics|predictive model|forecasting|experiment|a/b)\b", blob):
        return "Data Scientist"
    return "AI/ML"


def ddg_search(query: str, max_results: int = 8) -> list[Job]:
    url = f"https://duckduckgo.com/html/?q={quote_plus(query)}"
    headers = {"User-Agent": "Mozilla/5.0 job-digest-bot/1.0"}
    try:
        res = requests.get(url, headers=headers, timeout=20)
        res.raise_for_status()
    except Exception as exc:
        return [
            Job(
                company="Search fallback",
                role=query,
                location="Preferred locations",
                mode="Search",
                source="DuckDuckGo",
                link=url,
                snippet=f"Search failed: {exc}",
                score=1,
                next_action="Open search link manually",
            )
        ]

    soup = BeautifulSoup(res.text, "html.parser")
    jobs: list[Job] = []
    for item in soup.select(".result"):
        title_el = item.select_one(".result__a")
        snippet_el = item.select_one(".result__snippet")
        if not title_el:
            continue
        title = norm(title_el.get_text(" "))
        link = title_el.get("href") or url
        snippet = norm(snippet_el.get_text(" ") if snippet_el else "")
        blob = f"{title} {snippet}"
        if not is_allowed_location(blob):
            continue
        role = title
        company = infer_company(title, link)
        location = infer_location(blob)
        mode = "Remote" if is_remote(blob) else ("Hybrid" if "hybrid" in blob.lower() else "")
        score = score_job(role, company, location, snippet)
        if score >= 45:
            jobs.append(Job(company, role, location, mode, "Search", link, snippet[:350], score, "Verify listing and apply"))
        if len(jobs) >= max_results:
            break
    return jobs


def infer_company(title: str, link: str) -> str:
    parts = re.split(r"\s[-|–]\s|\sat\s", title, maxsplit=1, flags=re.I)
    if len(parts) > 1:
        return norm(parts[-1])[:80]
    host = urlparse(link).netloc.replace("www.", "")
    return host or "Unknown"


def infer_location(text: str) -> str:
    found = []
    for loc in ["Bengaluru", "Bangalore", "Pune", "Hyderabad", "Mumbai", "Remote", "Hybrid", "India"]:
        if loc.lower() in text.lower():
            found.append(loc)
    return ", ".join(dict.fromkeys(found)) or "Check listing"


def remoteok_jobs() -> list[Job]:
    try:
        res = requests.get("https://remoteok.com/api", headers={"User-Agent": "job-digest-bot/1.0"}, timeout=20)
        res.raise_for_status()
        data = res.json()[1:]
    except Exception:
        return []
    jobs = []
    for item in data[:50]:
        role = norm(item.get("position", ""))
        company = norm(item.get("company", ""))
        tags = " ".join(item.get("tags") or [])
        snippet = norm(f"{tags} {item.get('description', '')}")[:350]
        blob = f"{role} {company} {tags} {snippet}"
        if not re.search(r"\b(ai|ml|machine learning|data engineer|data scientist|python|llm|nlp|rag|genai)\b", blob, re.I):
            continue
        if re.search(r"\b(marketing|sales|recruiter|operations executive|customer support|copywriter|medical scribe)\b", blob, re.I):
            continue
        link = item.get("url") or f"https://remoteok.com/remote-jobs/{item.get('id')}"
        score = score_job(role, company, "Remote", blob)
        if score >= 60:
            jobs.append(Job(company, role, "Remote", "Remote", "RemoteOK", link, snippet, score, "Check contract/location constraints"))
    return jobs


def parse_naukri_html(html_text: str, max_per_search: int) -> list[Job]:
    soup = BeautifulSoup(html_text, "html.parser")
    jobs: list[Job] = []
    found = 0
    for link_el in soup.select('a[href*="job-listings"]'):
        link = link_el.get("href") or ""
        role = norm(link_el.get_text(" "))
        if not link or not role:
            continue
        card = link_el.find_parent(["article", "div", "li"])
        text = norm(card.get_text(" ") if card else role)
        if not text or not is_allowed_location(text):
            continue
        if re.search(r"\b(sales|marketing|recruiter|support|business development)\b", text, re.I):
            continue
        company = infer_company(role, link)
        company_el = card.select_one(".comp-name, .subTitle, .companyInfo") if card else None
        if company_el:
            company = norm(company_el.get_text(" "))[:80]
        location = infer_location(text)
        mode = "Remote" if is_remote(text) else ("Hybrid" if "hybrid" in text.lower() else "")
        score = score_job(role, company, location, text)
        if score >= 55:
            jobs.append(
                Job(
                    company=company,
                    role=role,
                    location=location,
                    mode=mode,
                    source="Naukri",
                    link=link,
                    snippet=text[:350],
                    score=score,
                    next_action="Open Naukri listing and apply/reach recruiter",
                )
            )
            found += 1
        if found >= max_per_search:
            break
    return jobs


def naukri_jobs(max_per_search: int = 4) -> list[Job]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
        )
    }
    jobs: list[Job] = []
    urls: list[tuple[str, str]] = []
    for queries in JOB_GROUPS.values():
        for query in queries:
            slug = re.sub(r"[^a-z0-9]+", "-", query.lower()).strip("-")
            urls.append(
                (
                    query,
                    f"https://www.naukri.com/{slug}-jobs"
                    f"?k={quote_plus(query)}"
                    "&l=Bangalore%2FBengaluru%2C%20Pune%2C%20Hyderabad%2C%20Mumbai%2C%20Remote"
                    "&experience=2&jobAge=7",
                )
            )
    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--no-sandbox"])
            page = browser.new_page(user_agent=headers["User-Agent"])
            for _, url in urls:
                try:
                    page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    page.wait_for_timeout(5000)
                    jobs.extend(parse_naukri_html(page.content(), max_per_search))
                except Exception:
                    continue
            browser.close()
    except Exception:
        for _, url in urls:
            try:
                res = requests.get(url, headers=headers, timeout=20)
                res.raise_for_status()
            except Exception:
                continue
            jobs.extend(parse_naukri_html(res.text, max_per_search))
    if not jobs:
        for query, url in urls:
            jobs.append(
                Job(
                    company="Naukri",
                    role=f"Naukri search: {query}",
                    location="Bengaluru, Pune, Hyderabad, Mumbai, Remote",
                    mode="Search",
                    source="Naukri",
                    link=url,
                    snippet=(
                        "Open this saved-preference Naukri search while logged in to see fresh matching jobs, "
                        "recruiter/application details, and Neo recommendations."
                    ),
                    score=58,
                    next_action="Open Naukri search and review new listings",
                )
            )
    return dedupe(jobs, limit=60)


def linkedin_jobs() -> list[Job]:
    jobs: list[Job] = []
    for queries in JOB_GROUPS.values():
        for query in queries:
            url = (
                "https://www.linkedin.com/jobs/search/"
                f"?f_TPR=r604800&f_WT=2%2C3&keywords={quote_plus(query)}&location=India"
            )
            jobs.append(
                Job(
                    company="LinkedIn",
                    role=f"LinkedIn search: {query}",
                    location="India / Remote / Hybrid",
                    mode="Search",
                    source="LinkedIn",
                    link=url,
                    snippet=(
                        "Open this LinkedIn search while logged in to see recent jobs, Easy Apply roles, "
                        "recruiter/company context, and saved-alert results."
                    ),
                    score=59,
                    next_action="Open LinkedIn search and review recent listings",
                )
            )
    return jobs


def collect_jobs() -> list[Job]:
    queries = [
        '"Generative AI Engineer" RAG Python 2-4 years Bangalore OR Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"AI/ML Engineer" Python LLM RAG Bangalore OR Pune OR Hyderabad OR Mumbai',
        '"Machine Learning Engineer" Python "2-4 years" Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"LLM Engineer" LangChain RAG Python Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"NLP Engineer" Python LLM Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"Data Engineer" Python SQL "2-4 years" Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"Data Engineer" ETL Python SQL Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"Data Engineer" Spark Python Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"Data Scientist" Python Machine Learning "2-4 years" Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"Data Scientist" NLP Python Bengaluru OR Pune OR Hyderabad OR Mumbai',
        '"Data Scientist" SQL Python Bengaluru OR Pune OR Hyderabad OR Mumbai',
        'site:foundit.in/job "Generative AI Engineer" "2-4" Hyderabad OR Pune OR Bengaluru',
        'site:foundit.in/job "Data Engineer" Python SQL Hyderabad OR Pune OR Bengaluru OR Mumbai',
        'site:foundit.in/job "Data Scientist" Python Hyderabad OR Pune OR Bengaluru OR Mumbai',
        'site:hirist.tech "Generative AI Engineer" RAG Python Bangalore Pune Hyderabad Mumbai',
        'site:hirist.tech "Data Engineer" Python SQL Bangalore Pune Hyderabad Mumbai',
        'site:hirist.tech "Data Scientist" Python Bangalore Pune Hyderabad Mumbai',
        'site:linkedin.com/jobs "Generative AI Engineer" RAG Python India',
        'site:linkedin.com/jobs "Data Engineer" Python SQL India',
        'site:linkedin.com/jobs "Data Scientist" Python India',
    ]
    jobs = []
    for q in queries:
        jobs.extend(ddg_search(q, max_results=5))
    jobs.extend(linkedin_jobs())
    jobs.extend(naukri_jobs())
    jobs.extend(remoteok_jobs())
    collected = balanced_jobs(dedupe(jobs, limit=100), per_group=5)
    if os.getenv("INCLUDE_FALLBACK_JOBS") == "1" and len(collected) < 15:
        collected = balanced_jobs(dedupe([*collected, *SEEDED_FALLBACK_JOBS], limit=100), per_group=5)
    return collected


def dedupe(jobs: Iterable[Job], limit: int | None = 15) -> list[Job]:
    seen = set()
    output = []
    for job in sorted(jobs, key=lambda j: j.score, reverse=True):
        key = job_key(job)
        if key in seen:
            continue
        seen.add(key)
        if job.score >= 45:
            output.append(job)
    return output[:limit] if limit is not None else output


def balanced_jobs(jobs: Iterable[Job], per_group: int = 5) -> list[Job]:
    buckets = {group: [] for group in JOB_GROUPS}
    extras = []
    for job in jobs:
        group = job_group(job)
        if group in buckets:
            buckets[group].append(job)
        else:
            extras.append(job)

    selected: list[Job] = []
    selected_keys = set()
    for group in JOB_GROUPS:
        group_selected = 0
        for source in ["LinkedIn", "Naukri"]:
            for job in [j for j in buckets[group] if j.source == source][:2]:
                key = job_key(job)
                if key not in selected_keys:
                    selected.append(job)
                    selected_keys.add(key)
                    group_selected += 1
                if group_selected >= per_group:
                    break
            if group_selected >= per_group:
                break
        for job in buckets[group][:per_group]:
            if group_selected >= per_group:
                break
            key = job_key(job)
            if key not in selected_keys:
                selected.append(job)
                selected_keys.add(key)
                group_selected += 1

    target_total = per_group * len(JOB_GROUPS)
    leftovers = [
        job
        for group in JOB_GROUPS
        for job in buckets[group][per_group:]
        if job_key(job) not in selected_keys
    ]
    leftovers.extend(job for job in extras if job_key(job) not in selected_keys)
    for job in sorted(leftovers, key=lambda j: j.score, reverse=True):
        if len(selected) >= target_total:
            break
        selected.append(job)
        selected_keys.add(job_key(job))
    return selected


def job_key(job: Job) -> str:
    link = job.link.strip().lower()
    if not is_live_search_link(job):
        link = re.sub(r"[?#].*$", "", link)
    return link or f"{job.company}|{job.role}|{job.location}".lower()


def load_seen(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    if not isinstance(data, dict):
        return {}
    return {str(k): v for k, v in data.items() if isinstance(v, dict)}


def save_seen(path: Path, seen: dict[str, dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(seen, indent=2, sort_keys=True), encoding="utf-8")


def filter_new_jobs(jobs: list[Job], seen: dict[str, dict[str, str]]) -> list[Job]:
    if os.getenv("INCLUDE_FALLBACK_JOBS") != "1":
        for seed in SEEDED_FALLBACK_JOBS:
            seen.setdefault(job_key(seed), {
                "company": seed.company,
                "role": seed.role,
                "first_seen": "seeded-before-dedup",
            })
    return [job for job in jobs if is_live_search_link(job) or job_key(job) not in seen]


def remember_jobs(jobs: list[Job], seen: dict[str, dict[str, str]], now: dt.datetime) -> dict[str, dict[str, str]]:
    stamp = now.strftime("%Y-%m-%dT%H:%M:%SZ")
    for job in jobs:
        if is_live_search_link(job):
            continue
        seen[job_key(job)] = {
            "company": job.company,
            "role": job.role,
            "location": job.location,
            "source": job.source,
            "link": job.link,
            "first_seen": seen.get(job_key(job), {}).get("first_seen", stamp),
            "last_sent": stamp,
        }
    # Keep the cache bounded.
    return dict(list(seen.items())[-500:])


def is_live_search_link(job: Job) -> bool:
    return job.source in {"LinkedIn", "Naukri"} and job.role.startswith(f"{job.source} search:")


def build_workbook(jobs: list[Job], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Tracker"
    headers = [
        "Status",
        "Priority",
        "Category",
        "Score",
        "Company",
        "Role",
        "Location",
        "Mode",
        "Source",
        "Apply Link",
        "Notes",
        "Next Action",
        "Applied Date",
        "Follow-up Date",
    ]
    ws.append(headers)
    for idx, job in enumerate(jobs, start=1):
        ws.append([
            "To Review",
            idx,
            job_group(job),
            job.score,
            job.company,
            job.role,
            job.location,
            job.mode,
            job.source,
            job.link,
            job.snippet,
            job.next_action,
            "",
            "",
        ])
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)
    widths = [14, 10, 18, 8, 24, 42, 26, 14, 14, 60, 60, 30, 15, 15]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if jobs:
        tab = Table(displayName="JobsTracker", ref=f"A1:N{len(jobs)+1}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)
    ws.freeze_panes = "A2"

    msg = wb.create_sheet("Templates")
    msg["A1"] = "Recruiter Message"
    msg["A1"].font = Font(bold=True, size=14)
    msg["A2"] = (
        "Hi [Name],\n\n"
        "I am Vraj Patel, an AI/ML Software Engineer with 2.5 years of experience building production AI solutions, "
        "including AWS Textract OCR pipelines, RAG chatbots, Neo4j knowledge graphs, Python microservices, data pipelines, "
        "and MLOps workflows.\n\n"
        "I saw the [Role] opening at [Company]. My experience with Python, LLM/RAG systems, LangChain/LangGraph, AWS, "
        "Neo4j, Docker, and production AI workflows looks closely aligned. I am available to join immediately and open "
        "to Bangalore, Pune, Hyderabad, Mumbai, or remote/hybrid roles.\n\n"
        "Could you please consider my profile for this role?\n\nRegards,\nVraj Patel"
    )
    msg["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    msg.column_dimensions["A"].width = 120
    msg.row_dimensions[2].height = 220

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_email_body(jobs: list[Job], total_found: int = 0, skipped_seen: int = 0) -> str:
    today = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    if not jobs:
        return (
            f"AI/ML Job Digest - {today}\n\n"
            "No new public matches were found in this run.\n\n"
            f"Checked matches: {total_found}\n"
            f"Skipped already-sent matches: {skipped_seen}\n\n"
            "The bot will keep checking every 2 hours and will send fresh matches when it finds them."
        )
    lines = [
        f"AI/ML Job Digest - {today}",
        "",
        "Top matches:",
        "",
    ]
    for idx, job in enumerate(jobs[:12], start=1):
        lines.extend([
            f"{idx}. {job.company} - {job.role}",
            f"Category: {job_group(job)}",
            f"Location/mode: {job.location} / {job.mode or 'Check listing'}",
            f"Score: {job.score}",
            f"Link: {job.link}",
            f"Why: {job.snippet[:240]}",
            f"Next action: {job.next_action}",
            "",
        ])
    lines.extend([
        "Top recruiter message:",
        "",
        "Hi [Name],",
        "",
        "I am Vraj Patel, an AI/ML Software Engineer with 2.5 years of experience building production AI solutions including AWS Textract OCR pipelines, RAG chatbots, Neo4j knowledge graphs, Python microservices, data pipelines, and MLOps workflows.",
        "",
        "I saw the [Role] opening at [Company]. My experience with Python, LLM/RAG systems, LangChain/LangGraph, AWS, Neo4j, Docker, and production AI workflows looks closely aligned. I am available to join immediately and open to Bangalore, Pune, Hyderabad, Mumbai, or remote/hybrid roles.",
        "",
        "Could you please consider my profile for this role?",
        "",
        "Regards,",
        "Vraj Patel",
    ])
    return "\n".join(lines)


def send_email(subject: str, body: str, attachment: Path | None = None) -> bool:
    host = os.getenv("SMTP_HOST") or "smtp.gmail.com"
    port = int(os.getenv("SMTP_PORT") or "587")
    username = os.getenv("SMTP_USERNAME") or SMTP_ACCOUNT
    password = os.getenv("SMTP_PASSWORD")
    mail_to = os.getenv("MAIL_TO") or SMTP_ACCOUNT
    mail_from = os.getenv("MAIL_FROM") or username or SMTP_ACCOUNT
    if not all([host, username, password, mail_to, mail_from]):
        missing = [
            name
            for name, value in {
                "SMTP_HOST": host,
                "SMTP_USERNAME": username,
                "SMTP_PASSWORD": password,
                "MAIL_TO": mail_to,
                "MAIL_FROM": mail_from,
            }.items()
            if not value
        ]
        message = f"SMTP secrets are not configured; missing: {', '.join(missing)}"
        if os.getenv("REQUIRE_EMAIL") == "1" or os.getenv("GITHUB_ACTIONS") == "true":
            raise RuntimeError(message)
        print(message + "; skipping email send.", file=sys.stderr)
        return False

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = mail_from
    msg["To"] = mail_to
    msg.set_content(body)
    if attachment and attachment.exists():
        data = attachment.read_bytes()
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment.name,
        )
    with smtplib.SMTP(host, port, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(username, password)
        smtp.send_message(msg)
    return True


def main() -> int:
    now = dt.datetime.now(dt.UTC)
    seen_path = Path(os.getenv("SEEN_JOBS_FILE", ".cache/seen_jobs.json"))
    seen = load_seen(seen_path)
    all_jobs = collect_jobs()
    jobs = filter_new_jobs(all_jobs, seen)
    skipped_seen = max(0, len(all_jobs) - len(jobs))
    output = Path("outputs") / f"vraj_job_digest_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    build_workbook(jobs, output)
    body = build_email_body(jobs, total_found=len(all_jobs), skipped_seen=skipped_seen)
    status = "New matches" if jobs else "No new matches"
    subject = f"AI/ML Job Digest - {status} - {now.strftime('%Y-%m-%d %H:%M UTC')}"
    print(f"Found {len(all_jobs)} jobs; prepared {len(jobs)} new jobs; skipped {skipped_seen} already seen.")
    for idx, job in enumerate(jobs[:8], start=1):
        print(f"{idx}. {job.company} - {job.role} [{job.score}] {job.link}")
    print(f"Excel tracker: {output}")
    try:
        if send_email(subject, body, output):
            seen = remember_jobs(jobs, seen, now)
            save_seen(seen_path, seen)
            print("Email sent successfully.")
        else:
            print("Email not sent because SMTP secrets are missing.")
    except Exception as exc:
        print(f"EMAIL_SEND_FAILED: {type(exc).__name__}: {exc}", file=sys.stderr)
        raise
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
